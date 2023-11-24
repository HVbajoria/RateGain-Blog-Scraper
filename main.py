# importing libraries
from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from bs4 import BeautifulSoup
import re
import requests
import pandas as pd
import streamlit as st
import time
from openpyxl.styles import PatternFill
from languages import languages

key = st.secrets['TRANSLATOR_KEY']
endpoint = st.secrets['TRANSLATOR_ENDPOINT']

from azure.ai.textanalytics import TextAnalyticsClient
from azure.core.credentials import AzureKeyCredential

# Authenticate the client using your key and endpoint 
def authenticate_client():
    ta_credential = AzureKeyCredential(key)
    text_analytics_client = TextAnalyticsClient(
            endpoint=endpoint, 
            credential=ta_credential)
    return text_analytics_client

client = authenticate_client()

# Translating the text to the desired language of the user
def translator(text, to_lang):
    import uuid
    path = '/translate?api-version=3.0'
    params = '&to='+to_lang
    constructed_url = endpoint + path + params

    headers = {
        'Ocp-Apim-Subscription-Key': key,
        'Ocp-Apim-Subscription-Region': st.secrets['TRANSLATOR_REGION'],
        'Content-type': 'application/json',
        'X-ClientTraceId': str(uuid.uuid4())
    }

    # You can pass more than one object in body.
    body = [{
        'text' : text
    }]
    request = requests.post(constructed_url, headers=headers, json=body)
    response = request.json()
    print(response)
    return(response[0]['translations'][0]['text'])

# extracting the maximum number of pages
def extract_max_page_number(base_url):
    options = webdriver.ChromeOptions()
    options.add_argument('headless')

    try:
        driver = webdriver.Chrome(options=options)
        driver.get(base_url)
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        pagination_div = soup.find('div', class_='pagination col-xs-12')

        # Check if pagination div exists where the page numbers are present
        if pagination_div:
            links = pagination_div.find_all('a', class_='page-numbers')
            # Extract numerical values from the links using regular expressions
            page_numbers = [int(re.search(r'\d+', link['href']).group()) for link in links]

            # Find the maximum page number
            max_page_number = max(page_numbers)
            return max_page_number
        else:
            print("Pagination div not found.")
            return None
    except WebDriverException as e:
        print(f"Error: {e}")
        return None
    finally:
        try:
            driver.quit()  # Ensure the WebDriver is closed even if an exception occurs
        except:
            pass

# Extracing the title, date, image url, and likes count from the blog posts on a page
def extract_blog_data(url, to_lang):
    options = webdriver.ChromeOptions()
    options.add_argument('headless')

    try:
        driver = webdriver.Chrome(options=options)
        driver.get(url)
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        # To store the extracted data
        blog_data = {
            'Title': [],
            'Date': [],
            'Image URL': [],
            'Likes Count': []
        }

        # Getting the division where all the blog posts are present
        blogs_div = soup.find('div', class_='blog-items')

        # Check if blog posts are present
        if blogs_div is None:
            print("No blog posts found.")
            print(url)
            return extract_blog_data(url)
        else:
            # Extract all the blog posts from the division
            blog_posts = blogs_div.find_all('article', class_='blog-item')
            print(f"Found {len(blog_posts)} blog posts.")

            # Extract the title, date, image url, and likes count from each blog post
            for post in blog_posts:
                title = post.find('h6').find('a').text
                title=translator(title, to_lang=list(languages.keys())[list(languages.values()).index(to_lang)])
                title = title.replace('â€™', '\'').strip()
                date = post.find('div', class_='bd-item').find('span').text
                date=translator(date, to_lang=list(languages.keys())[list(languages.values()).index(to_lang)])
                Image_URL = post.find('div', class_='img')
                if Image_URL:
                    Image_URL = Image_URL.find('a')['data-bg']
                else:
                    Image_URL = ""

                # Extract only the numeric part of the likes count
                Likes_Count_text = post.find('a', class_='zilla-likes').find('span').text
                Likes_Count = int(''.join(filter(str.isdigit, Likes_Count_text)))
                Likes_Count=translator(Likes_Count, to_lang=list(languages.keys())[list(languages.values()).index(to_lang)])
                # Append the extracted data to the dictionary
                blog_data['Title'].append(title)
                blog_data['Date'].append(date)
                blog_data['Image URL'].append(Image_URL)
                blog_data['Likes Count'].append(Likes_Count)

        return blog_data
    except WebDriverException as e:
        print(f"Error: {e}")
        return None
    finally:
        try:
            driver.quit()  # Ensure the WebDriver is closed even if an exception occurs
        except:
            pass

# Extracting the urls of all the pagination pages
def extract_pagination_urls(base_url, max_pages):
    pagination_urls = [f"{base_url}/page/{i}/" for i in range(2, max_pages + 1)]
    return pagination_urls

# Saving the extracted data to a csv file
# Function to save data to an Excel file with title row highlighted
def save_to_excel(data, file_path):
    try:
        df = pd.DataFrame(data)

        # Create an Excel writer with Pandas
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            # Get the xlsxwriter workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Create a fill object with the color you want to use
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Iterate through all columns and set the fill color for each cell in the first row (title row)
            for col_num, value in enumerate(df.columns.values):
                cell = worksheet.cell(row=1, column=col_num + 1)
                cell.fill = fill

    except Exception as e:
        st.error(f"Error saving to Excel: {e}")
        st.stop()

# Streamlit application
def main():
    # Set page configuration
    st.set_page_config(
        page_title="Blog Data Extraction",
        page_icon="📚",
        layout="centered",
    )

    hide_default_format = """ 
        <style> 
        #MainMenu {visibility: show; } 
        footer {visibility: hidden;} 
        </style> 
        """ 

    # Page header with logo and title
    st.markdown(hide_default_format, unsafe_allow_html=True) 

    def gradient_text(text, color1, color2):
        gradient_css = f"""
            background: -webkit-linear-gradient(left, {color1}, {color2});
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            font-weight: bold;
            font-size: 42px;
            """
        return f'<span style="{gradient_css}">{text}</span>'

    color1 = "#DC8772"
    color2 = "#FF9933"
    text = "Blog Data Extractor"
  
    left_co, cent_co,last_co = st.columns(3)
    with cent_co:
        st.image("logo.png", width=300)

    styled_text = gradient_text(text, color1, color2)
    st.write(f"<div style='text-align: center;'>{styled_text}</div>", unsafe_allow_html=True)
    st.markdown(
        """
        *Click the button below to get blog details and download the Excel file.*

        *Choose language to translate title, date and like count*
        """,
        unsafe_allow_html=True,
    )

    # To allow the user to select the language
    if 'selected_language' not in st.session_state:
        st.session_state.selected_language = 'English'
    to_lang = st.session_state.selected_language
    to_lang = st.selectbox("Select the language", languages.values(), index=list(languages.values()).index(st.session_state.selected_language))
    
    # Button to trigger data extraction and download
    if st.button("Get Blog Details"):
        # Use st.spinner to display a loading spinner while processing
        with st.spinner("Fetching and processing data..."):
            time.sleep(2)  # Simulate delay for demonstration purposes

            base_url = 'https://rategain.com/blog'
            max_pages = extract_max_page_number(base_url)

            if max_pages is not None:
                st.info(f"Number of pages: {max_pages}")
            else:
                max_pages = 1
                st.warning("Unable to determine the number of pages.")

            pagination_urls = extract_pagination_urls(base_url, max_pages)
            all_data = {
                'Title': [],
                'Date': [],
                'Image URL': [],
                'Likes Count': []
            }

            # Extract data from the main page
            main_page_data = extract_blog_data(base_url, to_lang)
            if main_page_data:
                all_data['Title'].extend(main_page_data['Title'])
                all_data['Date'].extend(main_page_data['Date'])
                all_data['Image URL'].extend(main_page_data['Image URL'])
                all_data['Likes Count'].extend(main_page_data['Likes Count'])

            # Extract data from pagination pages
            for url in pagination_urls:
                page_data = extract_blog_data(url, to_lang)
                if page_data:
                    all_data['Title'].extend(page_data['Title'])
                    all_data['Date'].extend(page_data['Date'])
                    all_data['Image URL'].extend(page_data['Image URL'])
                    all_data['Likes Count'].extend(page_data['Likes Count'])


            # Save the data to an Excel file
            save_to_excel(all_data, 'blog_data.xlsx')
            with open("blog_data.xlsx", "rb") as template_file:
                template_byte = template_file.read()

                # Provide a download button for the Excel file
                st.download_button(
                    label="Download Blog Data Excel File",
                    data=template_byte,
                    file_name='blog_data.xlsx',
                    key='excel-download'
                )

# Run the Streamlit app
if __name__ == "__main__":
    main()
